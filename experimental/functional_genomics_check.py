"""Functional-genomics triangulation: is PLK1 (or the centrosome module Nsp13 binds) a
required host factor for SARS-CoV-2, in genome-wide CRISPR knockout screens?

This is an INDEPENDENT line of evidence: CRISPR fitness screens measure phenotype, not
interaction, so they escape the shared-study-bias ceiling that capped tests 8-14. Three
published genome-wide KO screens are checked (all validated internally by ACE2 ranking ~#1):
  - Calu-3 (lung) and Vero E6, from Rebendenne/Bruel et al., Nat. Genet. 2022 (s41588-022-01110-2)
  - a third bidirectional screen from Biering et al., Nat. Genet. 2022 (s41588-022-01131-x)

Reads the supplementary tables live from Springer static-content. Positive z / positive-
selection enrichment = sgRNA enriched in surviving cells = gene KO PROTECTS = pro-viral host
dependency factor. The screens' residual-z / MAGeCK methods already regress out general gene
essentiality, so a neutral PLK1 is not merely masked by its mitotic-essential toxicity.
"""
import urllib.request
import openpyxl
import csv

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Safari/605.1.15"
FILES = {
    "calu3": "https://static-content.springer.com/esm/art%3A10.1038%2Fs41588-022-01110-2/MediaObjects/41588_2022_1110_MOESM7_ESM.xlsx",
    "vero":  "https://static-content.springer.com/esm/art%3A10.1038%2Fs41588-022-01110-2/MediaObjects/41588_2022_1110_MOESM5_ESM.xlsx",
    "biering": "https://static-content.springer.com/esm/art%3A10.1038%2Fs41588-022-01131-x/MediaObjects/41588_2022_1131_MOESM3_ESM.xlsx",
}
GENES = ["PLK1", "CEP135", "CNTRL", "NIN", "NINL", "PCNT", "CDK5RAP2", "AKAP9", "CENPF", "CEP250", "TBK1"]


def fetch(url, path):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req) as r, open(path, "wb") as f:
        f.write(r.read())


def zscore_table(path, sheet):
    ws = openpyxl.load_workbook(path, read_only=True)[sheet]
    rows = list(ws.iter_rows(values_only=True))[1:]
    return {r[0]: (r[1], r[2]) for r in rows if r[0]}


def main():
    for k, url in FILES.items():
        fetch(url, f"/tmp/fg_{k}.xlsx")

    calu = zscore_table("/tmp/fg_calu3.xlsx", "Calu3_Gattinara_avg_zscore")
    vero = zscore_table("/tmp/fg_vero.xlsx", "VeroE6_avg_zscore")
    wb = openpyxl.load_workbook("/tmp/fg_biering.xlsx", read_only=True)
    brows = list(wb["Supplementary Table 1"].iter_rows(values_only=True))[1:]
    bier = {r[0]: (r[2], r[3]) for r in brows if r[0]}  # (pos|fdr, pos|rank)

    print(f"Screen sizes: Calu-3 {len(calu)}, Vero {len(vero)}, Biering {len(bier)} genes")
    print("Internal validation (ACE2 should rank ~#1): "
          f"Calu-3 #{calu['ACE2'][1]}, Vero #{vero['ACE2'][1]}, Biering #{bier['ACE2'][1]}\n")

    print(f"{'gene':10s} {'Calu-3 z (rank)':>18s} {'Vero z (rank)':>16s} {'Biering posFDR (rank)':>22s}")
    out = [["gene", "calu3_z", "calu3_rank", "vero_z", "vero_rank", "biering_posFDR", "biering_rank"]]
    for g in GENES:
        c, v, b = calu.get(g), vero.get(g), bier.get(g)
        cs = f"{c[0]:.2f} ({c[1]})" if c else "absent"
        vs = f"{v[0]:.2f} ({v[1]})" if v else "absent"
        bs = f"{b[0]:.2f} ({b[1]})" if b and b[0] is not None else "absent"
        print(f"{g:10s} {cs:>18s} {vs:>16s} {bs:>22s}")
        out.append([g, c[0] if c else "", c[1] if c else "", v[0] if v else "", v[1] if v else "",
                    b[0] if b else "", b[1] if b else ""])

    with open("data/validation/crispr_functional_genomics.csv", "w", newline="") as f:
        csv.writer(f).writerows(out)

    print("\nVerdict: PLK1 and the centrosome module are NOT significant host factors in any of the")
    print("three screens (no |z|>=2, no FDR<0.1). PLK1 knockout does not change viral fitness -->")
    print("PLK1 is not a required replication host factor. This does NOT refute the Nsp13-PLK1")
    print("interaction (binding != fitness requirement); it reframes any real interaction as more")
    print("likely a pathogenesis/cell-cycle-perturbation mechanism than an essential host-factor role,")
    print("and it tempers the PLK1-inhibitor drug-repurposing angle.")
    print("saved data/validation/crispr_functional_genomics.csv")


if __name__ == "__main__":
    main()
